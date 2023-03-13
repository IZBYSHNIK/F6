from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side, Alignment, PatternFill
from openpyxl.styles.numbers import BUILTIN_FORMATS
import json
import os
import datetime
import calendar


class Student:

    def __init__(self, fio, sick_days=None, absence_days=None):
        if not sick_days:
            sick_days = {}
        if not absence_days:
            absence_days = {}
        self.__fio = self.chech_fio(fio)
        self.__sick_days = sick_days
        self.__absence_days = absence_days
        self.marks = {}

    def __str__(self):
        return self.create_shorts_fio(self.fio)

    @classmethod
    def chech_fio(cls, fio):
        if cls.is_valud_fio(fio):
            return fio.title()
        raise ValueError('ФИО должно состоять только из букв и быть из 3 частей, каждая из которых не менее 2 символов')

    @staticmethod
    def is_valud_fio(fio):
        return isinstance(fio.replace(' ', ''), str) and sum([item.isalpha() and len(item) >= 2 for item in fio.split()]) == 3


    @staticmethod
    def chech_day(day):
        if isinstance(day, int) and 1 <= day <= 31:
            return day
        raise ValueError('День должен быть целым числом, от 1 до 31')

    @staticmethod
    def chech_hours(hours):
        if isinstance(hours, int) and 1 <= hours <= 10:
            return hours
        raise ValueError('Часы должены быть записаны целым числом, от 1 до 10')

    @property
    def fio(self):
        return self.__fio

    @fio.setter
    def fio(self, value):
        self.__fio = self.chech_fio(value)

    @staticmethod
    def create_shorts_fio(fio: str):
        last_name, first_name, middle_name = fio.title().split()
        return f'{last_name} {first_name[0]}. {middle_name[0]}.'

    @property
    def sick_days(self):
        return self.__sick_days

    @property
    def absence_days(self):
        return self.__absence_days

    def add_sick_day(self, day, hours):
        day = self.chech_day(day)
        hours = self.chech_hours(hours)
        self.__sick_days[day] = hours

    def add_absence_day(self, day, hours):
        day = self.chech_day(day)
        hours = self.chech_hours(hours)
        self.__absence_days[day] = hours

    def get_statistic_for_student(self):
        return {
            'FIO': self.fio,
            "Sick_days": (self.sick_days, sum(self.sick_days.values())),
            "Absence_days": (self.absence_days, sum(self.absence_days.values())),
        }

    def add_day(self, number_day, hours, type_day='a'):
        if type_day == 's':
            self.add_sick_day(number_day, hours)
        elif type_day == 'a':
            self.add_absence_day(number_day, hours)
        else:
            raise ValueError("Неверно указан тип дня")

    # def add_mark(self, id_items, day):
    #     self.marks.append((id_items, day))


class ManagerStudents:
    CLASS_STUDENT = Student
    MONTHS = (
        'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
        'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
    )
    HAPPY_DAYS = {
        10: [4, ],
        1: [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, ],
        12: [30, 31, ],
        2: [23, ],
        3: [8, ],
    }

    def __init__(self, period,  user=None, days=None, students=None):
        self.user = user
        self.period = period
        self.days = days if days else self.generate_work_days()
        self.__students = [] if not students else students
        self.date = datetime.datetime.now().timetuple()
        self.parametrs = {}
        self.couples = {}
        self.is_archive = False

    def __getitem__(self, indx):
        return self.students[indx]

    def __setitem__(self, indx, student):
        self.__students[indx] = student


    def generate_work_days(self):
        work_days = {}
        for week in calendar.monthcalendar(year=self.period[1], month=self.period[0]):
            for day in range(7):
                if day > 4 or week[day] in self.HAPPY_DAYS.get(self.period[0], []):
                    continue
                elif week[day] != 0 and week[day]:
                    work_days[week[day]] = 0
        return work_days

    def off_day(self, day):
        if day in self.days:

            for st in self.students:
                if st.sick_days.get(day):
                    del st.sick_days[day]
                elif st.absence_days.get(day):
                    del st.absence_days[day]

            del self.days[day]

    def on_day(self, day, hours=0):
        self.days[day] = hours

    def add_hours_by_day(self, day, hours):
        if day in self.days and 0 <= hours <= 10:
            self.days[day] = hours

    @property
    def students(self):

        return self.__students

    def add_student(self, student):
        self.__students.append(student)

    def remove_student_id(self, id_s):
        del self.__students[id_s]

    def __delitem__(self, key):
        self.remove_student_id(key)

    def remove_student_obj(self, obj):
        self.__students.remove(obj)

    def save_students(self, file_name='students.json', path_file=None):
        """
        Сохраняет базу данных о студентах
        """

        if path_file is None:
            path_file = self.user.path
        date = self.serialization()
        if not os.path.isdir(self.user.path):
            os.mkdir(self.user.path)
        with open(os.path.join(path_file, file_name), 'w') as f:
            json.dump(date, f)


    def load_students(self, file_name='students.json', user_path=None):
        """Считывает данные из БД и возвращает из них словарь"""
        if user_path is None:
            user_path = self.user.path
        with open(os.path.join(user_path, file_name), 'r') as f:
            data = json.load(f)
        return data


    def serialization(self):
        """Преобразует данные для сохранения в Json"""

        liststudents, psc, R = self.__encode()
        data = {
            "Period": self.period,
            "Liststudents": liststudents,
            "Couples": self.couples,
            'Days': self.days,
            "Parametrs": self.parametrs,
            'PSC': psc,
            'Rang': R,
        }
        return data


    @staticmethod
    def crate_eternal_iter(data, indx=0):
        """Создает из итерируемого объекта бесконечный итератор"""
        while True:
            yield data[indx]
            indx += 1
            if indx == len(data):
                indx = 0

    @staticmethod
    def get_number_by_chars(number):
        """Преобразует число в буквенное представление. Имеет два режима если передать целое число - просто вернет
        его буквенное представление, но при передаче картежа вернет представление с разделяющей точкой"""
        excel_col_name = lambda n: '' if n <= 0 else excel_col_name((n - 1) // 26) + chr((n - 1) % 26 + ord('A'))
        if isinstance(number, tuple):
            a, b = number[0], number[-1]
            return excel_col_name(a) + '.' + excel_col_name(b)

        return excel_col_name(number)

    @staticmethod
    def get_chars_by_number(chars: str):
        """Преобразует буквенное представление в число. Имеет два режима если передать буквенное представление без точки - просто вернет
                его числовое представление, но при передаче буквенного представление с разделяющей точкой вернет картеж из целых чисел"""
        excel_col_num = lambda a: 0 if a == '' else 1 + ord(a[-1]) - ord('A') + 26 * excel_col_num(a[:-1])
        if '.' in chars:
            a, b = chars.split('.')
            return excel_col_num(a), excel_col_num(b)
        return excel_col_num(chars)


    @staticmethod
    def share_data(data: str, sep=3):
        """Разделяет строковые данные на части равные по длине sep"""
        result = []
        for i in range(0, len(data), sep):
            result.append(data[i:i + sep])
        return result

    def crate_encryption_table(self, data, _psc=None):
        """Создает позиционную таблицу сопоставления прежнему значению, со смещенеием на UNICODE-значение символа из пороля"""
        def get_range(chars):
            r = len(chars)
            rangs = {
                'R1': (lambda x: 1 <= x <= 26, (1, 26), 1),
                'R2': (lambda x: 27 <= x <= 702, (27, 702), 2),
                'R3': (lambda x: 703 <= x <= 18278, (703, 18278), 3),
                'R4': (lambda x: 18279 <= x <= 475254, (18279, 475254), 4),
                'R5': (lambda x: 475255 <= x <= 11881376, (475255, 11881376), 5),
            }
            for k, v in rangs.items():
                if v[0](r):
                    return v[1:]
            return None

        table = {}
        password = self.user.password
        p = self.crate_eternal_iter(password)
        uid = self.crate_eternal_iter(self.user.user_id)
        password_sum = sum([ord(i) for i in password])

        if not _psc:
            chars = set(data)
            R = get_range(chars)
            p_s = [data.rindex(char) for char in chars]
            _psc = [
                self.get_number_by_chars(
                    (p_s[i] + len(p_s) + ord(next(p)) + password_sum, ord(data[p_s[i]]) + ord(next(uid))))
                for i in range(len(p_s))]

            cells = iter(range(1, R[0][-1]))
            for i in chars:
                table[i] = self.get_number_by_chars(next(cells)).rjust(R[-1], '_')
        else:

            psc = [(self.get_chars_by_number(_psc[i])[0] - (len(_psc) + ord(next(p)) + password_sum),
                    chr(self.get_chars_by_number(_psc[i])[-1] - ord(next(uid)))) for i in range(len(_psc))]
            R = get_range(psc)

            data = self.share_data(data, R[-1])
            chars = [data[i[0]] for i in psc]
            for c, p in zip(chars, psc):
                table[c] = p[-1]

        return table, _psc, R[-1]

    def save_f6(self, file_name='f6.xlsx'):
        BASE_COORDS = (0, 0)

        CELLS_INIT = {
            self.get_number_by_chars(9 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): self.user.parametrs.get('specialization', ''),
            self.get_number_by_chars(21 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): self.user.parametrs.get('group', ''),
            self.get_number_by_chars(24 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): 'Кл.руководитель ' + (self.user.create_shorts_fio(self.user.parametrs.get('teamleader')) if self.user.parametrs.get('teamleader') else ''),
            self.get_number_by_chars(24 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]): "Староста группы " + (self.user.create_shorts_fio(self.user.parametrs.get('offical_name')) if self.user.parametrs.get('offical_name') else ''),
            self.get_number_by_chars(35+BASE_COORDS[0])+str(3+BASE_COORDS[1]): 'Из них',
            self.get_number_by_chars(34 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): f'за {self.MONTHS[self.period[0]-1]} {str(self.period[1])}',
            self.get_number_by_chars(1 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1]): '№',
            self.get_number_by_chars(36+BASE_COORDS[0])+str(4+BASE_COORDS[1]): 'по неув',
            self.get_number_by_chars(35+BASE_COORDS[0])+str(4+BASE_COORDS[1]): 'по ув',
            self.get_number_by_chars(34+BASE_COORDS[0])+str(3+BASE_COORDS[1]): 'Итого',
            self.get_number_by_chars(2+BASE_COORDS[0])+str(3+BASE_COORDS[1]): 'ФИО',
            self.get_number_by_chars(18+BASE_COORDS[0])+str(2+BASE_COORDS[1]): 'Группа',
            self.get_number_by_chars(3+BASE_COORDS[0])+str(2+BASE_COORDS[1]): 'Специальность',
            self.get_number_by_chars(1+BASE_COORDS[0])+str(1+BASE_COORDS[1]): 'ВЕДОМОСТЬ УЧЁТА ЧАСОВ, ПРОПУЩЕННЫХ СТУДЕНТАМИ',
            self.get_number_by_chars(3 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): 'Чел.Час',
            self.get_number_by_chars(6 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): '=',
            self.get_number_by_chars(3 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]): 'Посещ.об',
            self.get_number_by_chars(6 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]): '=',
            self.get_number_by_chars(12 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): 'Посящаемость Кач.',
            self.get_number_by_chars(17 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): '=',
            self.get_number_by_chars(12 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]): 'Прогул 1 студ.',
            self.get_number_by_chars(17 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]): '=',
            self.get_number_by_chars(29 + BASE_COORDS[0]) + str(35 + BASE_COORDS[1]): 'Итого',
        }
        CELLS_MERGE = [
            f'{self.get_number_by_chars(3+BASE_COORDS[0])+str(2+BASE_COORDS[1])}:{self.get_number_by_chars(8+BASE_COORDS[0])+str(2+BASE_COORDS[1])}',
            f'{self.get_number_by_chars(3+BASE_COORDS[0])+str(37+BASE_COORDS[1])}:{self.get_number_by_chars(5+BASE_COORDS[0])+str(37+BASE_COORDS[1])}',
            f'{self.get_number_by_chars(3+BASE_COORDS[0])+str(38+BASE_COORDS[1])}:{self.get_number_by_chars(5+BASE_COORDS[0])+str(38+BASE_COORDS[1])}',
            f'{self.get_number_by_chars(18+BASE_COORDS[0])+str(2+BASE_COORDS[1])}:{self.get_number_by_chars(20+BASE_COORDS[0])+str(2+BASE_COORDS[1])}',
            f'{self.get_number_by_chars(2+BASE_COORDS[0])+str(3+BASE_COORDS[1])}:{self.get_number_by_chars(2+BASE_COORDS[0])+str(4+BASE_COORDS[1])}',
            f'{self.get_number_by_chars(35+BASE_COORDS[0])+str(3+BASE_COORDS[1])}:{self.get_number_by_chars(36+BASE_COORDS[0])+str(3+BASE_COORDS[1])}',
            f'{self.get_number_by_chars(12+BASE_COORDS[0])+str(37+BASE_COORDS[1])}:{self.get_number_by_chars(16+BASE_COORDS[0])+str(37+BASE_COORDS[1])}',
            f'{self.get_number_by_chars(12+BASE_COORDS[0])+str(38+BASE_COORDS[1])}:{self.get_number_by_chars(16+BASE_COORDS[0])+str(38+BASE_COORDS[1])}',
            f'{self.get_number_by_chars(1+BASE_COORDS[0]) + str(3 + BASE_COORDS[1])}:{self.get_number_by_chars(1 + BASE_COORDS[0]) + str(4 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(29+BASE_COORDS[0])+str(35+BASE_COORDS[1])}:{self.get_number_by_chars(33+BASE_COORDS[0])+str(35+BASE_COORDS[1])}',
            f'{self.get_number_by_chars(1 + BASE_COORDS[0]) + str(1 + BASE_COORDS[1])}:{self.get_number_by_chars(36 + BASE_COORDS[0]) + str(1 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(34 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}:{self.get_number_by_chars(36 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(9 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}:{self.get_number_by_chars(15 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(21 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}:{self.get_number_by_chars(26 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(7 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}:{self.get_number_by_chars(10 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(18 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}:{self.get_number_by_chars(21 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(7 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}:{self.get_number_by_chars(10 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(18 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}:{self.get_number_by_chars(21 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(24 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}:{self.get_number_by_chars(33 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(24 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}:{self.get_number_by_chars(33 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}',

        ]
        SET_WIDTH_CELLS = {
            self.get_number_by_chars(1+BASE_COORDS[0]): 3,
            self.get_number_by_chars(2+BASE_COORDS[0]): 30,
            self.get_number_by_chars(34+BASE_COORDS[0]): 8,
            self.get_number_by_chars(35+BASE_COORDS[0]): 8,
            self.get_number_by_chars(36+BASE_COORDS[0]): 8,


        }

        statistics = self.get_statistics()
        if statistics:
            CELLS_INIT[self.get_number_by_chars(34 + BASE_COORDS[0]) + str(35 + BASE_COORDS[1])] = statistics.get('all_absence')
            CELLS_INIT[self.get_number_by_chars(35 + BASE_COORDS[0]) + str(35 + BASE_COORDS[1])] = statistics.get('sack_days_hours')
            CELLS_INIT[self.get_number_by_chars(36 + BASE_COORDS[0]) + str(35 + BASE_COORDS[1])] = statistics.get('absence_days_hours')
            CELLS_INIT[self.get_number_by_chars(34 + BASE_COORDS[0]) + str(4 + BASE_COORDS[1])] = statistics.get('all_hours')
            CELLS_INIT[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])] = statistics.get('man_hours')
            CELLS_INIT[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])] = statistics.get('total_attendance')
            CELLS_INIT[self.get_number_by_chars(18 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])] = statistics.get('quality_attendance')
            CELLS_INIT[self.get_number_by_chars(18 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])] = statistics.get('absences_by_student')


        wb = Workbook()
        ws = wb.active

        patternfull = PatternFill(fgColor='D9D9D9', fill_type='solid')
        font = Font(name='Calibri', size=16, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
                    color='FF000000')
        font_sick_days = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='single', strike=False,
                    color='FF000000')
        font_absence_days = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False,
                    color='FF000000')
        fill_border_style = Border(Side(border_style="thin", color="000000"), Side(border_style="thin", color="000000"),
                                   Side(border_style="thin", color="000000"), Side(border_style="thin", color="000000"))


        ws[self.get_number_by_chars(2+BASE_COORDS[0])+str(3+BASE_COORDS[1])].border = fill_border_style
        ws[self.get_number_by_chars(1+BASE_COORDS[0])+str(1+BASE_COORDS[1])].font = font
        ws[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])].number_format = BUILTIN_FORMATS[10]
        ws[self.get_number_by_chars(18 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])].number_format = BUILTIN_FORMATS[10]
        ws[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])].number_format = BUILTIN_FORMATS[1]
        ws[self.get_number_by_chars(18 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])].number_format = BUILTIN_FORMATS[1]

        for name_call in (
                self.get_number_by_chars(34 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]),
                self.get_number_by_chars(35 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]),
                self.get_number_by_chars(34 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]),
                self.get_number_by_chars(35 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]),
        ):
            ws[name_call].border = Border(bottom=Side(style='thin'))

        #Настройка ширины столбцов и нумерация дней


        for i in range(3, 34):
            name_cell = self.get_number_by_chars(i+BASE_COORDS[0])+str(4+BASE_COORDS[1])

            ws[name_cell] = i - 2
            ws.column_dimensions[self.get_number_by_chars(i+BASE_COORDS[0])].width = 3.5

            if i-2 in self.days:
                ws[self.get_number_by_chars(i+BASE_COORDS[0])+str(3+BASE_COORDS[1])] = self.days.get(i-2)

            else:
                ws[name_cell].fill = patternfull
                ws[self.get_number_by_chars(i + BASE_COORDS[0]) + str(3 + BASE_COORDS[1])] = '✖'

        for name, value in CELLS_INIT.items():
            ws[name] = value
            ws[name].alignment = Alignment(horizontal='center')

        for i in CELLS_MERGE:
            ws.merge_cells(i)

        for key, value in SET_WIDTH_CELLS.items():
            ws.column_dimensions[key].width = value

        students = self.students[:30]

        column = self.get_number_by_chars(2+BASE_COORDS[0])

        for indx, student in enumerate(students):
            row = str(5 + indx + BASE_COORDS[1])
            fio, s_d, a_d = student.get_statistic_for_student().values()
            ws[column+row] = self.CLASS_STUDENT.create_shorts_fio(fio)
            ws[self.get_number_by_chars(self.get_chars_by_number(column)-1)+row] = indx + 1
            for c in range(3, 34):
                if c-2 in a_d[0]:
                    ws[self.get_number_by_chars(c+BASE_COORDS[0]) + row] = a_d[0].get(c-2, '')
                    ws[self.get_number_by_chars(c+BASE_COORDS[0]) + row].font = font_absence_days

                elif c-2 in s_d[0]:
                    ws[self.get_number_by_chars(c+BASE_COORDS[0]) + row] = s_d[0].get(c - 2, '')
                    ws[self.get_number_by_chars(c+BASE_COORDS[0]) + row].font = font_sick_days

            ws[self.get_number_by_chars(35 + BASE_COORDS[0]) + row] = s_d[1] if s_d[1] else ''
            ws[self.get_number_by_chars(36 + BASE_COORDS[0]) + row] = a_d[1] if a_d[1] else ''

        for i in range(1, 37):
            for j in range(3, 36):
                ws[self.get_number_by_chars(i+BASE_COORDS[0]) + str(j+BASE_COORDS[1])].border = fill_border_style
                ws[self.get_number_by_chars(i+BASE_COORDS[0]) + str(j+BASE_COORDS[1])].alignment = Alignment(horizontal='center')
                if 3 <= i <= 33:
                    if i-2 not in self.days:
                        ws[self.get_number_by_chars(i + BASE_COORDS[0]) + str(j)].fill = patternfull

        wb.save(file_name)
        return file_name



    def save_f6_marks(self, file_name='f6marks.xlsx'):
        BASE_COORDS = (0, 0)

        CELLS_INIT = {
            self.get_number_by_chars(4 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): self.user.parametrs.get('specialization', ''),
            self.get_number_by_chars(9 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): self.user.parametrs.get('group', ''),
            self.get_number_by_chars(10 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1]): 'Кл.руководитель ' + (self.user.create_shorts_fio(self.user.parametrs.get('teamleader')) if self.user.parametrs.get('teamleader') else ''),
            self.get_number_by_chars(10 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): "Староста группы " + (self.user.create_shorts_fio(self.user.parametrs.get('offical_name')) if self.user.parametrs.get('offical_name') else ''),
            self.get_number_by_chars(12 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): f'за {self.MONTHS[self.period[0]-1]} {str(self.period[1])}',
            self.get_number_by_chars(1 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1]): '№',
            self.get_number_by_chars(2+BASE_COORDS[0])+str(3+BASE_COORDS[1]): 'ФИО',
            self.get_number_by_chars(8+BASE_COORDS[0])+str(2+BASE_COORDS[1]): 'Группа',
            self.get_number_by_chars(3+BASE_COORDS[0])+str(2+BASE_COORDS[1]): 'Специальность',
            self.get_number_by_chars(1+BASE_COORDS[0])+str(1+BASE_COORDS[1]): 'ВЕДОМОСТЬ УЧЁТА ЧАСОВ, УСПИВАЕМОСТИ СТУДЕНТАМИ',
            self.get_number_by_chars(2 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1]): 'Кол-во студ. им-х 2',
            self.get_number_by_chars(2 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): 'Кол-во студ. им-х 5',
            self.get_number_by_chars(2 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]): 'Кол-во студ. им-х одну 3',
            self.get_number_by_chars(2 + BASE_COORDS[0]) + str(39 + BASE_COORDS[1]): 'Кол-во студ. им-х 4 и 5',
            self.get_number_by_chars(5 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1]): 'Успеваемость общая',
            self.get_number_by_chars(5 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): 'Успеваемость качественная',

        }
        CELLS_MERGE = [
            f'{self.get_number_by_chars(2+BASE_COORDS[0])+str(3+BASE_COORDS[1])}:{self.get_number_by_chars(2+BASE_COORDS[0])+str(4+BASE_COORDS[1])}',
            f'{self.get_number_by_chars(1+BASE_COORDS[0]) + str(3 + BASE_COORDS[1])}:{self.get_number_by_chars(1 + BASE_COORDS[0]) + str(4 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(1 + BASE_COORDS[0]) + str(1 + BASE_COORDS[1])}:{self.get_number_by_chars(13 + BASE_COORDS[0]) + str(1 + BASE_COORDS[1])}',

        ]
        SET_WIDTH_CELLS = {
            self.get_number_by_chars(1+BASE_COORDS[0]): 3,
            self.get_number_by_chars(2+BASE_COORDS[0]): 30,
        }
        statistics = self.get_statistics_marks()
        if statistics:
            CELLS_INIT[self.get_number_by_chars(3 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1])] = statistics.get('heaving_2')
            CELLS_INIT[self.get_number_by_chars(3 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])] = statistics.get('heaving_5')
            CELLS_INIT[self.get_number_by_chars(3 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])] = statistics.get('heaving_one_3')
            CELLS_INIT[self.get_number_by_chars(3 + BASE_COORDS[0]) + str(39 + BASE_COORDS[1])] = statistics.get('heaving_4_and_5')
            CELLS_INIT[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1])] = statistics.get('total_academic_performance')
            CELLS_INIT[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])] = statistics.get('quality_academic_performance')

        wb = Workbook()
        ws = wb.active

        font = Font(name='Calibri', size=16, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
                    color='FF000000')
        fill_border_style = Border(Side(border_style="thin", color="000000"), Side(border_style="thin", color="000000"),
                                   Side(border_style="thin", color="000000"), Side(border_style="thin", color="000000"))


        ws[self.get_number_by_chars(2+BASE_COORDS[0])+str(3+BASE_COORDS[1])].border = fill_border_style
        ws[self.get_number_by_chars(1+BASE_COORDS[0])+str(1+BASE_COORDS[1])].font = font
        ws[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])].number_format = BUILTIN_FORMATS[10]
        ws[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1])].number_format = BUILTIN_FORMATS[10]


        for name_call in (
                self.get_number_by_chars(12 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1]),
                self.get_number_by_chars(13 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]),
                self.get_number_by_chars(12 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]),
                self.get_number_by_chars(13 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1]),
        ):
            ws[name_call].border = Border(bottom=Side(style='thin'))

        #Настройка ширины столбцов и нумерация дней


        for i in range(3, 14):
            name_call1 = self.get_number_by_chars(i+BASE_COORDS[0])+str(3+BASE_COORDS[1])
            name_cell2 = self.get_number_by_chars(i+BASE_COORDS[0])+str(4+BASE_COORDS[1])
            ws[name_call1] = self.couples.get(str(i-2))[0] if self.couples.get(str(i-2)) else ''
            ws[name_cell2] = self.couples.get(str(i-2))[1] if self.couples.get(str(i-2)) else ''
            ws.column_dimensions[self.get_number_by_chars(i+BASE_COORDS[0])].width = 15
            


        for name, value in CELLS_INIT.items():
            ws[name] = value
            ws[name].alignment = Alignment(horizontal='center')

        for i in CELLS_MERGE:
            ws.merge_cells(i)

        for key, value in SET_WIDTH_CELLS.items():
            ws.column_dimensions[key].width = value

        students = self.students[:30]

        column = self.get_number_by_chars(2+BASE_COORDS[0])

        for indx, student in enumerate(students):
            row = str(5 + indx + BASE_COORDS[1])
            fio, s_d, a_d = student.get_statistic_for_student().values()
            ws[column+row] = self.CLASS_STUDENT.create_shorts_fio(fio)
            ws[self.get_number_by_chars(self.get_chars_by_number(column)-1)+row] = indx + 1

            for i in student.marks:
                ws[self.get_number_by_chars(i+2) + row] = student.marks.get(i, '')



        for i in range(1, 14):
            for j in range(3, 35):
                ws[self.get_number_by_chars(i+BASE_COORDS[0]) + str(j+BASE_COORDS[1])].border = fill_border_style
                ws[self.get_number_by_chars(i+BASE_COORDS[0]) + str(j+BASE_COORDS[1])].alignment = Alignment(horizontal='center')
                ws.row_dimensions[j].height = 15

        wb.save(file_name)
        return file_name

    def __encode(self):
        data = self.__convert_students()
        table, psc, R = self.crate_encryption_table(data)
        result = ''
        for i in data:
            result += table.get(i, '')
        return result, psc, R

    def decode(self, data, psc):
        table, psc, R = self.crate_encryption_table(data, psc)
        result = ''
        for i in self.share_data(data, R):
            result += table.get(i, '')
        return result

    def __convert_students(self, sep='!'):
        result = ['#']
        for st in self.__students:
            result.extend([st.fio, ''.join([str(i).rjust(2, '0') + str(st.sick_days[i]).rjust(2, '0') for i in st.sick_days]), ''.join([str(i).rjust(2, '0') + str(st.absence_days[i]).rjust(2, '0') for i in st.absence_days]), ''.join([str(k).rjust(2, '0') + str(st.marks[k]).rjust(2, '0') for k in st.marks])])
        result.append('#')
        return f'{sep}'.join(result)

    @staticmethod
    def convert_str_to_list(date, sep='!'):

        def convert_data(data: str):
            result = {}
            for i in range(len(data) // 4):
                day = int(data[i*4] + data[i*4+1])
                hours = int(data[i*4+2] + data[i*4+3])

                result[day] = hours
            return result

        date = date[2:-2].split(sep)
        students = []
        for i in range(len(date)//4):
            fio = date[i*4]
            s_d = date[i*4+1]
            a_d = date[i*4+2]
            marks = date[i * 4 + 3]
            s = Student(fio, convert_data(s_d), convert_data(a_d))

            result = {}
            for i in range(0, len(marks), 4):
                raw = marks[i:i + 4]
                result[int(raw[0:2])] = int(raw[2:])
            s.marks = result
            students.append(s)

        return students

    @staticmethod
    def load_manager_students(user, data=None, file_name='students.json'):
        if data is None:
            user_path = user.path
            with open(os.path.join(user_path, file_name), 'r') as f:
                data = json.load(f)

        new_obj = ManagerStudents(data['Period'], user, {int(k): int(v) for k, v in data['Days'].items()})

        students = new_obj.convert_str_to_list(new_obj.decode(data['Liststudents'], data['PSC']))
        students = list(sorted(students, key=lambda item: item.fio))
        for student in students:
            new_obj.add_student(student)

        new_obj.parametrs = data['Parametrs']
        new_obj.couples = data.get('Couples', {})
        return new_obj

    def add_day(self, student, number_day, hours, type_day='a'):
        type_day = type_day.lower()
        if type_day == 's':
            student.add_day(number_day, hours, type_day)
        elif type_day == 'a':
            student.add_day(number_day, hours, type_day)
        else:
            raise ValueError("Неверно указан тип дня")

    def del_day(self, student, number_day):
        if number_day in student.sick_days:
            del student.sick_days[number_day]
        if number_day in student.absence_days:
            del student.absence_days[number_day]

    def set_period(self, month, year):
        self.period = (int(month), int(year))
        self.days = self.generate_work_days()

    def get_statistics(self):
        all_hours = sum(list(self.days.values()))
        sack_days_hours = 0
        absence_days_hours = 0
        for student in self.students:
            statistic = student.get_statistic_for_student()
            sack_days_hours += statistic['Sick_days'][1]
            absence_days_hours += statistic['Absence_days'][1]

        all_absence = sack_days_hours + absence_days_hours
        man_hours = len(self.students) * all_hours
        statisitcs = {
            'all_hours': all_hours,
            'sack_days_hours': sack_days_hours,
            'absence_days_hours': absence_days_hours,
            'all_absence': all_absence,
            'man_hours': man_hours,
        }
        if man_hours > 0:
            total_attendance = (man_hours-absence_days_hours)/man_hours
            quality_attendance = (man_hours - all_absence)/man_hours
            absences_by_student = all_absence/len(self.students)

            statisitcs['total_attendance'] = total_attendance
            statisitcs["quality_attendance"] = quality_attendance
            statisitcs['absences_by_student'] = absences_by_student
        else:
            statisitcs['total_attendance'] = 0
            statisitcs["quality_attendance"] = 0
            statisitcs['absences_by_student'] = 0

        return statisitcs

    def get_statistics_marks(self):
        statisitcs = {
            'heaving_2': 0,
            "heaving_5": 0,
            'heaving_one_3': 0,
            "heaving_4_and_5": 0,
            'is_ready': 0,
            'no_att': 0,
        }
        for i in self.students:
            statisitcs['is_ready'] = 1
            marks = [st for st in i.marks.values()]
            if len(marks) >= len(self.couples):
                if 2 in marks:
                    statisitcs['heaving_2'] += 1
                elif all(map(lambda x: x == 5, marks)):
                    statisitcs['heaving_5'] += 1
                elif marks.count(3) == 1:
                    statisitcs['heaving_one_3'] += 1
                else:
                    statisitcs['heaving_4_and_5'] += 1
        count_student = len(self.students)
        if count_student > 0:
            statisitcs['total_academic_performance'] = ((count_student - statisitcs['heaving_2']) / count_student)
            statisitcs['quality_academic_performance'] = (statisitcs['heaving_4_and_5'] / count_student)
        else:
            statisitcs['total_academic_performance'] = 0
            statisitcs['quality_academic_performance'] = 0
        return statisitcs

    def push_archive(self, *args, **kargs):
        if not os.path.exists(os.path.join(self.user.path, 'archive')):
            os.makedirs(os.path.join(self.user.path, "archive"))
        file_path, file_name = self.create_archive_file_path(*args, **kargs)


        self.replace_file(os.path.join(self.user.path, 'students.json'),
                   os.path.join(self.user.path, 'archive', 'students.json'))
        self.rename_file(os.path.join(self.user.path, 'archive', 'students.json'), file_path)



    def create_archive_file_path(self, file_name=None, auto=False):
        if not file_name:
            file_name = self.create_archive_file_name()

        if os.path.exists(os.path.join(self.user.path, 'archive', file_name)) and not auto:
            raise FileExistsError

        id_copy = 1
        while os.path.exists(os.path.join(self.user.path, 'archive', file_name)):
            file_name = self.create_archive_file_name(id_copy)
            id_copy += 1

        return os.path.join(self.user.path, 'archive', file_name), file_name

    def create_archive_file_name(self, level=''):
        if level:
            level = ' (' + str(level) + ')'
        return str(self.period[1]) + '_' + str(self.period[0]).rjust(2, '0') + self.MONTHS[
                self.period[0] - 1] + level + '.json'



    def create_new_table(self):
        if 1 <= self.period[0] < 12:
            month = self.period[0] + 1
            year = self.period[1]
        else:
            month = 1
            year = self.period[1] + 1

        self.set_period(month, year)
        self.couples.clear()
        for s in self.students:
            s.sick_days.clear()
            s.absence_days.clear()
            s.marks.clear()
        self.save_students()

    def replace_file(self, path1, path2):
        os.replace(path1, path2)

    def rename_file(self, path1, path2):
        os.rename(path1, path2)

    def init_archive(self):
        path_archive = os.path.join(self.user.path, 'archive')
        if not os.path.exists(os.path.join(path_archive)):
            os.makedirs(os.path.join(path_archive))
        walk = tuple(os.walk(path_archive))
        self.path_archive_files = walk[0][0]
        self.name_files = sorted([i for i in walk[0][-1] if i.endswith('.json') and len(i) >= 15], reverse=True)
        return self.path_archive_files, self.name_files

    def load_archive_file(self, file_name, file_path=None):
        if file_name in self.name_files:
            return self.load_manager_students(self.user, self.load_students(file_name=file_name, user_path=file_path), file_name)
        return self

    def clear_marks(self):
        for i in range(len(self.students)):
            self.students[i].marks.clear()

    def clear_absences(self):
        for i in range(len(self.students)):
            self.students[i].sick_days.clear()
            self.students[i].absence_days.clear()

    def get_statistics_for_graph(self):
        statistics_resalt = {}
        for i in self.students:
            statistics = i.get_statistic_for_student()
            resalt = statistics['Sick_days'][1]+statistics['Absence_days'][1]
            if resalt > 0:
                statistics_resalt[statistics['FIO']] = resalt


        statistics_resalt = dict(sorted(statistics_resalt.items(), key=lambda item: item[1], reverse=True))
        return statistics_resalt

    def get_total_statistic_period(self):
        months = {
            #month: (посещ.качест, посещ.общая,)
        }
        path_archive_files, name_files = self.init_archive()
        for i in name_files:
            month = self.load_archive_file(i, path_archive_files)
            statistic = month.get_statistics()
            months[(month.period[1], month.period[0]-1)] = (statistic['sack_days_hours'], statistic['absence_days_hours'], statistic['all_absence'])

        months = dict(sorted(months.items(), key=lambda item: item))

        return months


    def del_file_archive(self, file_name):
        os.remove(os.path.join(self.path_archive_files, file_name))











if __name__ == '__main__':
    pass